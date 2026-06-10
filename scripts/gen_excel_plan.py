import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY = "1F3A5F"
DARK_BLUE = "2C5F8F"
MED_BLUE = "4472B0"
LIGHT_BLUE = "B8D0EA"
VERY_LIGHT_BLUE = "DCE6F3"
WHITE = "FFFFFF"
DARK_GRAY = "333333"
GRAY = "6B7280"
GREEN = "15803D"
AMBER = "D97706"
RED = "B91C1C"

title_font = Font(name="Microsoft YaHei", size=20, bold=True, color=NAVY)
subtitle_font = Font(name="Microsoft YaHei", size=11, color=GRAY, italic=True)
section_font = Font(name="Microsoft YaHei", size=12, bold=True, color=WHITE)
table_header_font = Font(name="Microsoft YaHei", size=11, bold=True, color=WHITE)
table_cell_font = Font(name="Microsoft YaHei", size=10, color=DARK_GRAY)
small_font = Font(name="Microsoft YaHei", size=9, color=GRAY, italic=True)
label_font = Font(name="Microsoft YaHei", size=11, bold=True, color=NAVY)
value_font = Font(name="Microsoft YaHei", size=11, color=DARK_GRAY)

navy_fill = PatternFill("solid", fgColor=NAVY)
dark_blue_fill = PatternFill("solid", fgColor=DARK_BLUE)
light_blue_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
very_light_blue_fill = PatternFill("solid", fgColor=VERY_LIGHT_BLUE)
white_fill = PatternFill("solid", fgColor=WHITE)

thin_border = Border(
    left=Side(style="thin", color=LIGHT_BLUE),
    right=Side(style="thin", color=LIGHT_BLUE),
    top=Side(style="thin", color=LIGHT_BLUE),
    bottom=Side(style="thin", color=LIGHT_BLUE),
)

center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)


def write_title(ws, row_num, title, subtitle=None, max_col=10):
    """写大标题，返回下一个可用行号"""
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=max_col)
    c = ws.cell(row=row_num, column=1, value=title)
    c.font = title_font
    c.alignment = Alignment(horizontal="left", vertical="center")
    if subtitle:
        ws.merge_cells(start_row=row_num + 1, start_column=1, end_row=row_num + 1, end_column=max_col)
        s = ws.cell(row=row_num + 1, column=1, value=subtitle)
        s.font = subtitle_font
    return row_num + (3 if subtitle else 2)


def write_table(ws, start_row, headers, rows, widths=None, header_fill=navy_fill):
    """写表格数据，返回下一个可用行号"""
    max_col = len(headers)
    # 表头行
    for c in range(1, max_col + 1):
        cell = ws.cell(row=start_row, column=c)
        cell.fill = header_fill
        cell.font = table_header_font
        cell.alignment = center_align
        cell.border = thin_border
    for c, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=c, value=h)
    # 数据行
    for ri, row in enumerate(rows):
        r = start_row + 1 + ri
        fill = very_light_blue_fill if ri % 2 == 1 else white_fill
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.font = table_cell_font
            cell.border = thin_border
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)
        ws.cell(row=r, column=1).alignment = left_align
        for c in range(2, max_col + 1):
            ws.cell(row=r, column=c).alignment = center_align
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    return start_row + len(rows) + 2


def write_section_label(ws, row_num, text, max_col=10, fill=navy_fill):
    """写一个跨列的深蓝色章节标题条，返回下一行号"""
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill = fill
        cell.font = section_font
        cell.alignment = center_align
        cell.border = thin_border
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=max_col)
    ws.cell(row=row_num, column=1, value=text)
    return row_num + 1


def write_label_value(ws, row_num, label, value, max_col=10):
    """写一行标签-值（分左右两列布局）"""
    # 左半：标签占2列，值占3列
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
    l1 = ws.cell(row=row_num, column=1, value=label)
    l1.font = label_font
    l1.fill = light_blue_fill
    l1.alignment = left_align
    l1.border = thin_border
    half = max_col // 2
    ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=half)
    v1 = ws.cell(row=row_num, column=3, value=value)
    v1.font = value_font
    v1.fill = white_fill
    v1.alignment = left_align
    v1.border = thin_border


def write_two_label_value(ws, row_num, label1, value1, label2, value2, max_col=10):
    """写一行，两列标签-值对"""
    half = max_col // 2
    # 左半
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
    l1 = ws.cell(row=row_num, column=1, value=label1)
    l1.font = label_font
    l1.fill = light_blue_fill
    l1.alignment = left_align
    l1.border = thin_border
    ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=half)
    v1 = ws.cell(row=row_num, column=3, value=value1)
    v1.font = value_font
    v1.fill = white_fill
    v1.alignment = left_align
    v1.border = thin_border
    # 右半
    ws.merge_cells(start_row=row_num, start_column=half + 1, end_row=row_num, end_column=half + 2)
    l2 = ws.cell(row=row_num, column=half + 1, value=label2)
    l2.font = label_font
    l2.fill = light_blue_fill
    l2.alignment = left_align
    l2.border = thin_border
    ws.merge_cells(start_row=row_num, start_column=half + 3, end_row=row_num, end_column=max_col)
    v2 = ws.cell(row=row_num, column=half + 3, value=value2)
    v2.font = value_font
    v2.fill = white_fill
    v2.alignment = left_align
    v2.border = thin_border


# ============ 创建工作簿 ============
wb = openpyxl.Workbook()
MAX_C = 10  # 默认列数

# ============ Sheet 1: 项目总览 ============
ws1 = wb.active
ws1.title = "01 项目总览"
ws1.sheet_view.showGridLines = False
for i, w in enumerate([16, 18, 28, 20, 14, 20, 12, 18, 12, 14], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

r = write_title(ws1, 1, "5D好车 B2B 二手车交易平台", "功能开发计划 · 高管汇报用", max_col=MAX_C)

info_items = [
    ("项目名称", "5D好车 B2B 二手车交易平台", "项目编号", "5D-CAR-PJT-001"),
    ("版本", "v1.0.0 功能正式上线版", "项目类型", "企业级 SaaS 应用"),
    ("发起人", "产品研发部", "项目经理", "张 伟"),
    ("技术架构", "Spring Boot 3.5 + PostgreSQL 16 + Redis 7 + RocketMQ 5",
     "技术负责人", "李 明"),
    ("后端接口", "96 个 RESTful API", "前端页面", "24 个 uni-app 页面"),
    ("数据库表", "31 张核心表", "功能模块", "13 个核心业务模块"),
    ("项目状态", "核心功能开发完成 · 后续优化推进中",
     "计划上线", "2026-06-15"),
    ("总预算", "¥ 1,850,000", "已投入", "¥ 1,420,000"),
]
r = write_section_label(ws1, r, "项目基本信息", max_col=MAX_C)
for item in info_items:
    write_two_label_value(ws1, r, item[0], item[1], item[2], item[3], max_col=MAX_C)
    r += 1
r += 1

# KPI 卡片
r = write_section_label(ws1, r, "项目核心指标", max_col=MAX_C)
kpi_cards = [
    ("后端 API 完成率", "~95%", "90+ / 96 接口"),
    ("前端页面完成率", "~88%", "24 / 27 页面"),
    ("数据库设计", "100%", "31 张核心表"),
    ("功能模块覆盖", "~93%", "13 / 13 模块"),
    ("核心 BUG 数", "0", "已全部清零"),
]
# 5 张卡片，每张占 2 列 × 3 行
for i, (label, value, sub) in enumerate(kpi_cards):
        col_start = i * 2 + 1
        # 第1行：大值
        ws1.merge_cells(start_row=r, start_column=col_start, end_row=r, end_column=col_start + 1)
        vc = ws1.cell(row=r, column=col_start, value=value)
        vc.font = Font(name="Microsoft YaHei", size=18, bold=True, color=GREEN)
        vc.alignment = center_align
        vc.fill = very_light_blue_fill
        vc.border = thin_border
        # 第2行：标签
        ws1.merge_cells(start_row=r + 1, start_column=col_start, end_row=r + 1, end_column=col_start + 1)
        lc = ws1.cell(row=r + 1, column=col_start, value=label)
        lc.font = Font(name="Microsoft YaHei", size=9, color=GRAY)
        lc.alignment = center_align
        lc.fill = very_light_blue_fill
        lc.border = thin_border
        # 第3行：子说明
        ws1.merge_cells(start_row=r + 2, start_column=col_start, end_row=r + 2, end_column=col_start + 1)
        sc = ws1.cell(row=r + 2, column=col_start, value=sub)
        sc.font = small_font
        sc.alignment = center_align
        sc.fill = very_light_blue_fill
        sc.border = thin_border
r += 4

# 项目目标
r = write_section_label(ws1, r, "项目目标与商业价值", max_col=MAX_C)
goals = [
    ("平台定位", "面向 B 端车商的全流程二手车交易平台，提供车源展示、交易撮合、信用体系、AI 服务"),
    ("商业目标", "上线 3 个月内注册车商 2,000+，日活跃车源 5,000+，月 GMV ¥ 5,000 万+"),
    ("核心价值", "车源透明化 · 交易标准化 · 风控自动化 · 获客智能化"),
    ("竞争优势", "全流程 B2B 交易闭环、AI 智能分析、保证金+信用双重风控、电子合同法律保障"),
    ("目标客户", "全国二手车车商、汽车经销集团、出口贸易商、汽车金融机构"),
    ("监管合规", "实名认证、交易留痕、电子合同备案、数据加密、符合《电子商务法》要求"),
]
for goal_lbl, goal_val in goals:
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    lc = ws1.cell(row=r, column=1, value=goal_lbl)
    lc.font = label_font
    lc.fill = light_blue_fill
    lc.alignment = left_align
    lc.border = thin_border
    ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=MAX_C)
    vc = ws1.cell(row=r, column=3, value=goal_val)
    vc.font = value_font
    vc.fill = white_fill
    vc.alignment = left_align
    vc.border = thin_border
    r += 1
r += 1

# 核心功能模块概览
r = write_section_label(ws1, r, "核心功能模块概览", max_col=MAX_C)
modules = [
    ("M01", "找车模块", "车源搜索、筛选、推荐、详情展示", "P0", "19 个功能点", "🟢 95%"),
    ("M02", "交易模块", "卖车/买车管理、订单流转、支付", "P0", "11 个功能点", "🟢 95%"),
    ("M03", "车源发布", "车辆信息录入、图片上传、定价", "P0", "6 个功能点", "🟢 95%"),
    ("M04", "用户中心", "个人信息、认证、信用体系", "P0", "10 个功能点", "🟢 98%"),
    ("M05", "消息系统", "站内消息、交易通知、系统公告", "P1", "10 个功能点", "🟢 90%"),
    ("M06", "AI 助理", "AI 行情分析、智能推荐、获客文案", "P1", "5 个功能点", "🟢 90%"),
    ("M07", "金融服务", "保证金管理、信用额度、账单明细", "P2", "4 个功能点", "🟢 90%"),
    ("M08", "出口业务", "出口车源标记、报关支持", "P2", "3 个功能点", "🟢 85%"),
    ("M09", "车行管理", "车行信息、员工管理、权限控制", "P1", "5 个功能点", "🟡 75%"),
    ("M10", "电子合同", "合同生成、签署、模板管理", "P1", "5 个功能点", "🟢 95%"),
    ("M11", "在线客服", "实时客服聊天、工单系统", "P2", "4 个功能点", "🟢 90%"),
    ("M12", "关注系统", "关注/取关卖家、关注列表", "P2", "3 个功能点", "🟢 95%"),
    ("M13", "会员体系", "会员等级、权益、续费", "P2", "4 个功能点", "🟢 90%"),
]
r = write_table(ws1, r,
                ["模块编号", "模块名称", "核心功能描述", "优先级", "功能点数", "完成状态"],
                modules, widths=[12, 16, 42, 10, 16, 12])

# ============ Sheet 2: 开发阶段 ============
ws2 = wb.create_sheet("02 开发阶段")
ws2.sheet_view.showGridLines = False
for i, w in enumerate([8, 20, 14, 14, 10, 40, 36, 16, 12], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

r = write_title(ws2, 1, "开发阶段与里程碑", "按阶段拆分的项目交付路径", max_col=9)

phase_data = [
    ("P0", "需求分析与架构设计", "2026-05-01", "2026-05-10", "2 周",
     "PRD、架构设计、数据库设计、API 规范",
     "✅ 架构评审通过 · 数据库方案确认", "产品+架构", "✅ 完成"),
    ("P1", "核心骨架与基础服务", "2026-05-11", "2026-05-21", "2 周",
     "脚手架、认证、基础 CRUD、消息中心",
     "✅ 基础服务联调通过 · 登录闭环", "后端团队", "✅ 完成"),
    ("P2", "车源与交易核心", "2026-05-22", "2026-06-05", "2 周",
     "车源发布/搜索/详情、订单、保证金、合同",
     "✅ 核心交易闭环 · 车源完整上架", "后端+前端", "✅ 完成"),
    ("P3", "拍卖与 AI 高级模块", "2026-06-06", "2026-06-10", "1 周",
     "拍卖竞价(10 接口)、AI 行情/获客/全网找车",
     "✅ 拍卖定时任务 · LLM 联调", "后端+AI", "✅ 完成"),
    ("P4", "金融与信用体系", "2026-06-11", "2026-06-18", "1 周",
     "保证金、充值/提现、信用额度、优惠券、会员",
     "✅ 金融模块联调 · 会员续费去重", "后端+前端", "✅ 完成"),
    ("P5", "车行、关注与消息", "2026-06-19", "2026-06-22", "1 周",
     "车行管理、成员、关注/取关、WebSocket 聊天",
     "✅ 实时聊天推送 · 客服工单闭环", "后端+前端", "✅ 完成"),
    ("P6", "联调测试与性能优化", "2026-06-23", "2026-06-27", "1 周",
     "全链路联调、压力测试、Redis 缓存、代码审查",
     "🔄 进行中 · 单接口 500 QPS 压测", "测试+后端", "🔄 进行中"),
    ("P7", "生产部署与灰度上线", "2026-06-28", "2026-07-03", "1 周",
     "生产部署、SSL 证书、域名备案、灰度 10%",
     "📅 计划 · 灰度测试无重大问题", "运维+后端", "📅 待启动"),
    ("P8", "正式运营与持续迭代", "2026-07-04", "2026-08-15", "6 周",
     "全面上线、运营数据跟踪、用户反馈、功能迭代",
     "📅 目标 · 月活 2,000+ · 月 GMV ¥ 5,000 万", "全团队+运营", "📅 待启动"),
]

r = write_table(ws2, r,
                ["阶段", "阶段名称", "开始", "结束", "工期",
                 "核心交付物", "关键里程碑", "负责团队", "状态"],
                phase_data, header_fill=dark_blue_fill)

# 里程碑时间线
r = write_section_label(ws2, r, "关键里程碑时间线", max_col=9, fill=navy_fill)
milestones = [
    ("MS-001", "2026-05-10", "架构评审通过", "技术架构确认、数据库设计完成", "架构委员会"),
    ("MS-002", "2026-05-21", "基础服务联调", "用户注册/登录/消息中心可用", "后端团队"),
    ("MS-003", "2026-06-05", "核心交易闭环", "车源发布→搜索→下单→合同", "产品+研发"),
    ("MS-004", "2026-06-10", "拍卖 + AI 上线", "实时竞价 + LLM 智能助理", "后端+AI"),
    ("MS-005", "2026-06-18", "会员与金融完成", "保证金 + 信用额度 + 优惠券 + 会员", "全团队"),
    ("MS-006", "2026-06-22", "前端核心页面完成", "车源/交易/合同/AI/消息/金融页面", "前端团队"),
    ("MS-007", "2026-06-27", "联调测试完成", "全链路测试 + 压力测试通过", "测试团队"),
    ("MS-008", "2026-07-03", "灰度上线完成", "生产部署 + 10% 灰度验证", "运维+后端"),
    ("MS-009", "2026-07-10", "正式全面上线", "100% 流量 + 运营推广", "全团队"),
    ("MS-010", "2026-08-15", "首期运营目标达成", "月活 2,000+ · 月 GMV ¥ 5,000 万", "运营+管理层"),
]
r = write_table(ws2, r,
                ["里程碑编号", "日期", "里程碑", "说明", "负责方"],
                milestones, widths=[14, 14, 22, 48, 18],
                header_fill=dark_blue_fill)

# ============ Sheet 3: 功能模块 ============
ws3 = wb.create_sheet("03 功能模块")
ws3.sheet_view.showGridLines = False
for i, w in enumerate([10, 16, 45, 10, 12, 12, 22, 14, 10], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

r = write_title(ws3, 1, "功能模块详细规划", "13 个核心业务模块 · 功能点与优先级梳理", max_col=9)

modules_detail = [
    ("M01", "找车模块", "车源搜索、筛选、排序、详情页、AI 车况、图片轮播",
     "P0", "19 个", 25, "李明/王芳", "—", "低"),
    ("M02", "交易模块", "订单创建、状态流转、支付保证金、确认收货、争议处理",
     "P0", "11 个", 30, "张伟/刘洋", "M01/M04", "中"),
    ("M03", "车源发布", "车辆信息录入、图片上传、价格设置、出口标记",
     "P0", "6 个", 18, "陈强/赵静", "M04", "低"),
    ("M04", "用户中心", "登录注册、实名认证、车商认证、信用等级、浏览记录",
     "P0", "10 个", 22, "李明/王芳", "—", "低"),
    ("M05", "消息系统", "系统消息、交易通知、活动通知、已读/未读、消息订阅",
     "P1", "10 个", 15, "孙磊/刘洋", "M04", "低"),
    ("M06", "AI 助理", "AI 行情、全网找车、自动拓客、获客文案、AI 分发",
     "P1", "5 个", 20, "周磊/李明", "M01/M04", "中"),
    ("M07", "金融服务", "保证金账户、充值/提现、信用额度申请、账单明细",
     "P2", "4 个", 18, "张伟/赵静", "M04", "高"),
    ("M08", "出口业务", "出口车源标记、报关信息、出口筛选、海外买家对接",
     "P2", "3 个", 12, "陈强/王芳", "M01/M03", "中"),
    ("M09", "车行管理", "车行信息、员工邀请/审批、成员列表、权限、业绩统计",
     "P1", "5 个", 20, "孙磊/刘洋", "M04", "中"),
    ("M10", "电子合同", "合同模板、合同生成、双方签署、查看/下载、合同归档",
     "P1", "5 个", 16, "张伟/赵静", "M02", "中"),
    ("M11", "在线客服", "客服会话发起、实时消息、常见问题、客服工单",
     "P2", "4 个", 14, "孙磊/刘洋", "M04/M05", "低"),
    ("M12", "关注系统", "关注/取消关注、关注列表、关注卖家车源动态",
     "P2", "3 个", 8, "陈强/王芳", "M01/M04", "低"),
    ("M13", "会员体系", "会员等级定义、会员权益、开通续费、到期提醒",
     "P2", "4 个", 10, "李明/赵静", "M04/M07", "低"),
]

r = write_table(ws3, r,
                ["模块编号", "模块名称", "业务描述", "优先级", "功能点数",
                 "估算人日", "负责人", "依赖模块", "风险等级"],
                modules_detail)

# ============ Sheet 4: 任务分解 WBS ============
ws4 = wb.create_sheet("04 任务分解")
ws4.sheet_view.showGridLines = False
for i, w in enumerate([8, 28, 12, 42, 22, 14, 14, 10, 10, 14], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

r = write_title(ws4, 1, "工作分解结构 (WBS)", "按任务维度 · 含责任人与时间安排", max_col=10)

wbs_tasks = [
    ("1.0", "项目启动与架构", "—", "立项、技术选型、架构设计、环境搭建",
     "张伟(PM)", "2026-05-01", "2026-05-10", 10, "P0", "✅ 100%"),
    ("1.1", "数据库设计", "架构", "31 张表结构、索引、约束、关系",
     "李明", "2026-05-01", "2026-05-07", 7, "P0", "✅ 100%"),
    ("1.2", "API 规范设计", "架构", "RESTful 接口、认证、错误码",
     "李明", "2026-05-03", "2026-05-10", 8, "P0", "✅ 100%"),
    ("2.0", "用户与认证服务", "M04", "登录注册、JWT 认证、权限控制",
     "李明", "2026-05-11", "2026-05-18", 8, "P0", "✅ 100%"),
    ("2.1", "注册/登录接口", "M04", "手机号+密码、JWT Token",
     "李明", "2026-05-11", "2026-05-14", 4, "P0", "✅ 100%"),
    ("2.2", "用户信息管理", "M04", "信息 CRUD、头像上传、实名认证",
     "李明", "2026-05-14", "2026-05-18", 5, "P0", "✅ 100%"),
    ("3.0", "车源核心模块", "M01/M03", "车源发布、搜索、详情、推荐",
     "陈强", "2026-05-15", "2026-06-02", 19, "P0", "✅ 100%"),
    ("3.1", "车源发布", "M03", "车辆信息录入、图片上传、定价、状态",
     "陈强", "2026-05-15", "2026-05-22", 8, "P0", "✅ 100%"),
    ("3.2", "车源搜索与列表", "M01", "多条件筛选、分页、排序、标签过滤",
     "陈强", "2026-05-20", "2026-05-28", 9, "P0", "✅ 100%"),
    ("3.3", "车源详情页", "M01", "详情数据聚合、图片轮播、参数展示",
     "陈强", "2026-05-25", "2026-06-02", 9, "P0", "✅ 100%"),
    ("3.4", "Redis 缓存层", "架构", "车源列表缓存、热点数据、自动刷新",
     "李明", "2026-05-28", "2026-06-03", 7, "P1", "✅ 100%"),
    ("4.0", "交易与订单", "M02", "订单创建、状态流转、保证金、争议",
     "张伟", "2026-05-22", "2026-06-10", 20, "P0", "✅ 100%"),
    ("4.1", "订单核心", "M02", "订单创建、状态变更、订单列表",
     "张伟", "2026-05-22", "2026-05-30", 9, "P0", "✅ 100%"),
    ("4.2", "支付与保证金", "M02", "保证金充值、订单锁定、退款",
     "张伟", "2026-05-27", "2026-06-05", 10, "P0", "✅ 100%"),
    ("4.3", "争议处理", "M02", "争议发起、平台仲裁、退款处理",
     "张伟", "2026-06-01", "2026-06-10", 10, "P1", "✅ 100%"),
    ("5.0", "电子合同", "M10", "合同模板、生成、签署、归档",
     "张伟", "2026-05-30", "2026-06-08", 10, "P1", "✅ 100%"),
    ("6.0", "消息与聊天", "M05/M11", "系统消息、WebSocket、客服工单",
     "孙磊", "2026-05-25", "2026-06-15", 22, "P1", "✅ 100%"),
    ("7.0", "AI 助理模块", "M06", "LLM 接入、行情分析、获客文案",
     "周磊", "2026-06-01", "2026-06-10", 10, "P1", "✅ 100%"),
    ("8.0", "拍卖系统", "拍卖", "拍卖列表、详情、出价、定时结算",
     "陈强", "2026-06-05", "2026-06-12", 8, "P1", "✅ 100%"),
    ("9.0", "金融与会员", "M07/M13", "保证金、信用额度、优惠券、会员",
     "张伟", "2026-06-10", "2026-06-18", 9, "P1", "✅ 100%"),
    ("10.0", "车行与关注", "M09/M12", "车行管理、成员、关注/取关",
     "孙磊", "2026-06-12", "2026-06-18", 7, "P1", "✅ 100%"),
    ("11.0", "前端开发", "全模块", "uni-app 页面、组件、交互、UI 适配",
     "王芳/刘洋/赵静", "2026-05-15", "2026-06-22", 39, "P0", "🟢 88%"),
    ("12.0", "测试与优化", "—", "联调测试、压力测试、代码审查、安全加固",
     "测试团队", "2026-06-20", "2026-06-27", 8, "P0", "🔄 进行中"),
    ("13.0", "部署与上线", "—", "生产部署、灰度上线、正式运营",
     "运维+全团队", "2026-06-28", "2026-07-10", 13, "P0", "📅 待启动"),
]

r = write_table(ws4, r,
                ["WBS", "任务名称", "所属模块", "任务描述", "负责人",
                 "开始日期", "结束日期", "工期(天)", "优先级", "完成状态"],
                wbs_tasks)

# ============ Sheet 5: 资源分配 ============
ws5 = wb.create_sheet("05 资源分配")
ws5.sheet_view.showGridLines = False
for i, w in enumerate([14, 16, 40, 32, 12, 10, 16, 12], 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

r = write_title(ws5, 1, "资源分配与团队配置", "人力资源 · 服务器资源 · 预算配置", max_col=8)

team_data = [
    ("张伟", "项目经理", "项目规划、进度跟踪、风险管理、跨团队协调",
     "敏捷管理/产品设计", 75, 2000, "¥ 150,000", "🔄 在职"),
    ("李明", "技术负责人", "架构设计、核心模块、代码审查、性能优化",
     "Spring Boot/PostgreSQL/Redis/架构", 90, 2200, "¥ 198,000", "🔄 在职"),
    ("陈强", "后端工程师", "车源模块、拍卖系统、关注系统、数据库层",
     "MyBatis-Plus/PostgreSQL/分布式锁", 85, 1800, "¥ 153,000", "🔄 在职"),
    ("张伟(后端)", "后端工程师", "交易模块、电子合同、金融服务、会员体系",
     "订单系统/支付/电子合同/账户体系", 85, 1800, "¥ 153,000", "🔄 在职"),
    ("孙磊", "后端工程师", "消息系统、WebSocket 聊天、客服工单",
     "RocketMQ/WebSocket/STOMP/实时通信", 80, 1800, "¥ 144,000", "🔄 在职"),
    ("周磊", "AI 工程师", "AI 助理、LLM 接入、行情分析、获客文案",
     "LLM/Prompt/RAG/推荐算法", 60, 2500, "¥ 150,000", "🔄 在职"),
    ("王芳", "前端工程师", "用户页面、车源页面、发布/详情页",
     "uni-app/Vue 3/uView/小程序", 70, 1600, "¥ 112,000", "🔄 在职"),
    ("刘洋", "前端工程师", "交易页面、合同页面、消息/聊天页面",
     "uni-app/Vue 3/组件设计/状态管理", 70, 1600, "¥ 112,000", "🔄 在职"),
    ("赵静", "前端工程师", "AI 页面、金融/会员页、拍卖/车行/优惠券",
     "uni-app/Vue 3/响应式/UI 优化", 65, 1600, "¥ 104,000", "🔄 在职"),
    ("吴昊", "测试工程师", "测试用例、联调、压力测试、Bug 跟踪",
     "JMeter/Postman/自动化测试", 50, 1400, "¥ 70,000", "🔄 在职"),
    ("郑爽", "UI/UX 设计师", "界面设计、Logo、图标、设计规范",
     "Figma/原型设计/UI 规范/交互设计", 30, 1800, "¥ 54,000", "🔄 在职"),
    ("钱进", "运维工程师", "服务器部署、CI/CD、监控告警、生产运维",
     "Linux/Docker/Nginx/Prometheus", 25, 2000, "¥ 50,000", "🔄 在职"),
    ("冯晓", "产品经理", "产品需求、原型、优先级管理、用户反馈",
     "Axure/需求分析/竞品调研/数据分析", 40, 2000, "¥ 80,000", "🔄 在职"),
    ("褚明", "安全顾问", "安全审计、渗透测试、合规检查、加密方案",
     "安全审计/OWASP/加密/合规", 10, 3000, "¥ 30,000", "📅 待启动"),
]

r = write_table(ws5, r,
                ["姓名", "角色", "主要职责", "技能/专长",
                 "投入人日", "¥/人日", "预估费用", "状态"],
                team_data)

# ============ Sheet 6: 风险评估 ============
ws6 = wb.create_sheet("06 风险评估")
ws6.sheet_view.showGridLines = False
for i, w in enumerate([6, 22, 48, 10, 10, 26, 12, 48, 10], 1):
    ws6.column_dimensions[get_column_letter(i)].width = w

r = write_title(ws6, 1, "项目风险评估与应对", "识别 · 评估 · 应对策略", max_col=9)

risks = [
    ("R01", "支付对接合规风险",
     "平台涉及保证金充值/提现等真实资金流动，需对接合规支付渠道并完成金融监管备案",
     "高", "中", "延迟上线/合规罚款",
     "张伟", "1. 提前对接第三方支付; 2. 与法务确认资金托管方案; 3. 准备备案资料",
     "🔴 高"),
    ("R02", "AI 服务不稳定",
     "LLM 服务依赖第三方 API，可能出现响应延迟、限流、内容审核问题",
     "中", "中", "AI 功能体验下降/运营投诉",
     "周磊", "1. 配置多 LLM 供应商兜底; 2. 实现本地缓存与重试; 3. 完善 Prompt 审核",
     "🟡 中"),
    ("R03", "车源数据造假",
     "平台上线后可能出现虚假车源、不实车况描述，影响平台可信度",
     "中", "高", "用户信任下降/平台口碑受损",
     "李明", "1. 车商实名认证+保证金制度; 2. 接入第三方车况检测报告; 3. 举报+人工审核",
     "🟡 中"),
    ("R04", "前端开发进度滞后",
     "拍卖页/车行页/优惠券页尚未完成，若前端开发不及时影响整体上线",
     "中", "中", "上线日期延后 1-2 周",
     "张伟(PM)", "1. 增加前端人力/加班; 2. 优先保障核心页面; 3. 后端 API 已就绪",
     "🟡 中"),
    ("R05", "数据库性能瓶颈",
     "车源搜索多条件组合查询、订单表增长较快，可能出现慢查询",
     "低", "中", "页面响应慢/并发能力不足",
     "李明", "1. 关键查询字段建立复合索引; 2. 引入 Redis 缓存热点车源; 3. 读写分离",
     "🟢 低"),
    ("R06", "WebSocket 连接稳定性",
     "实时聊天需保持长连接，网络不稳定或大规模并发可能导致断开",
     "低", "中", "聊天消息丢失/体验差",
     "孙磊", "1. 断线自动重连; 2. 消息本地存储+服务端 ACK; 3. 心跳保活",
     "🟢 低"),
    ("R07", "电子合同法律合规",
     "电子合同生成、签署、存储需符合《电子签名法》要求",
     "中", "高", "合同无效/法律纠纷",
     "张伟", "1. 对接可靠电子签名服务; 2. 合同模板法务审核; 3. 签署日志留痕",
     "🟡 中"),
    ("R08", "服务器安全与数据保护",
     "用户隐私数据(手机号、实名认证)需严格保护，防止数据泄露",
     "低", "高", "用户隐私泄露/监管处罚",
     "褚明+钱进", "1. 敏感字段加密存储; 2. HTTPS 全站加密; 3. 定期安全审计; 4. 访问日志",
     "🟡 中"),
    ("R09", "拍卖并发出价风险",
     "多个用户同时对同一拍品出价，需确保数据一致性、防止超价",
     "低", "中", "出价数据错乱/竞价纠纷",
     "陈强", "1. Redis 分布式锁+乐观锁; 2. 数据库事务隔离; 3. 实时价格广播校验",
     "🟢 低"),
    ("R10", "消息队列可靠性",
     "RocketMQ 消息可能丢失或重复消费，影响订单/通知等关键业务",
     "低", "中", "关键业务消息丢失/数据不一致",
     "孙磊", "1. 生产者确认机制; 2. 消息持久化; 3. 消费者幂等; 4. 死信队列告警",
     "🟢 低"),
]

r = write_table(ws6, r,
                ["编号", "风险描述", "详细说明",
                 "发生概率", "影响程度", "潜在影响",
                 "责任人", "应对策略", "风险等级"],
                risks)

# ============ Sheet 7: 成功指标 KPI ============
ws7 = wb.create_sheet("07 成功指标")
ws7.sheet_view.showGridLines = False
for i, w in enumerate([8, 24, 42, 10, 14, 14, 14, 14, 14], 1):
    ws7.column_dimensions[get_column_letter(i)].width = w

r = write_title(ws7, 1, "项目成功指标与 KPI", "量化目标 · 跟踪维度", max_col=9)

kpis = [
    ("BIZ-01", "月活跃车商数(MAU)", "注册并活跃的车商数量", "个",
     "500", "1,000", "2,000", "2026-09-30", "运营团队"),
    ("BIZ-02", "月新增车源数", "平台新增的二手车源数量", "个",
     "2,000", "5,000", "10,000", "2026-09-30", "运营团队"),
    ("BIZ-03", "月 GMV", "平台月交易总额", "元",
     "¥ 1,000,000", "¥ 3,000,000", "¥ 5,000,000", "2026-09-30", "财务+运营"),
    ("BIZ-04", "订单成交率", "成功交易订单数/总订单数", "%",
     "30%", "50%", "70%", "2026-09-30", "产品团队"),
    ("BIZ-05", "会员付费转化率", "付费会员/注册车商", "%",
     "5%", "10%", "15%", "2026-09-30", "运营团队"),
    ("BIZ-06", "AI 功能使用率", "使用 AI 功能的活跃用户占比", "%",
     "20%", "35%", "50%", "2026-09-30", "AI+产品"),
    ("TECH-01", "系统可用性 SLA", "服务正常运行时间占比", "%",
     "99.0%", "99.5%", "99.9%", "持续", "运维团队"),
    ("TECH-02", "首屏加载时间", "车源列表页面首次加载耗时", "秒",
     "≤3.0", "≤2.0", "≤1.5", "2026-07-15", "前端团队"),
    ("TECH-03", "API 平均响应时间", "后端接口 P95 响应时间", "ms",
     "≤500", "≤300", "≤200", "2026-07-15", "后端团队"),
    ("TECH-04", "接口并发能力", "系统 QPS 承载能力", "QPS",
     "200", "500", "1,000", "2026-07-15", "后端+运维"),
    ("TECH-05", "接口文档覆盖率", "后端接口文档与测试用例覆盖", "%",
     "80%", "90%", "100%", "2026-07-10", "测试团队"),
    ("USER-01", "用户满意度评分", "NPS 或满意度问卷得分", "分",
     "35", "50", "70", "2026-08-15", "产品团队"),
    ("USER-02", "客服工单处理时长", "客服工单从创建到关闭平均耗时", "小时",
     "≤24", "≤12", "≤6", "2026-07-15", "客服团队"),
    ("SEC-01", "重大安全事故数", "重大安全漏洞或数据泄露事件", "起",
     "0", "0", "0", "持续", "安全+运维"),
    ("FIN-01", "项目预算控制率", "实际费用/预算费用", "%",
     "≤110%", "≤105%", "≤100%", "2026-08-15", "财务+PM"),
]

r = write_table(ws7, r,
                ["编号", "指标名称", "指标说明", "单位",
                 "基线目标", "期望目标", "挑战目标", "达成日期", "责任方"],
                kpis)

# ============ Sheet 8: 交付物清单 ============
ws8 = wb.create_sheet("08 交付物清单")
ws8.sheet_view.showGridLines = False
for i, w in enumerate([8, 26, 44, 18, 16, 32, 14], 1):
    ws8.column_dimensions[get_column_letter(i)].width = w

r = write_title(ws8, 1, "项目交付物清单", "各阶段产出物与验收标准", max_col=7)

deliverables = [
    ("D-01", "产品需求文档(PRD)", "项目范围、功能需求、用户故事",
     "产品团队", "2026-05-10", "架构评审会评审通过", "✅ 已交付"),
    ("D-02", "系统架构设计文档", "后端架构图、前端架构图、数据流",
     "架构+后端", "2026-05-10", "架构评审会评审通过", "✅ 已交付"),
    ("D-03", "数据库设计文档", "31 张核心表 ER 图、字段说明、索引策略",
     "李明", "2026-05-10", "DBA 评审通过", "✅ 已交付"),
    ("D-04", "API 接口文档", "96 个 RESTful 接口、请求/响应规范",
     "后端团队", "2026-05-15", "前后端对接确认", "✅ 已交付"),
    ("D-05", "后端代码仓库", "Spring Boot 3.5 源码、单元测试",
     "后端团队", "2026-06-18", "代码 Review 通过", "✅ 已交付"),
    ("D-06", "前端代码仓库", "uni-app 3 源码、组件库、页面",
     "前端团队", "2026-06-22", "代码 Review 通过", "🟢 88%"),
    ("D-07", "数据库初始化脚本", "schema.sql + init.sql 初始化数据",
     "李明", "2026-05-12", "可重复执行、符合设计文档", "✅ 已交付"),
    ("D-08", "部署文档", "生产环境部署步骤、配置说明、Docker Compose",
     "钱进+李明", "2026-06-28", "部署验证通过", "🔄 进行中"),
    ("D-09", "运维手册", "服务器维护、故障排查、数据备份恢复",
     "钱进", "2026-07-05", "运维团队评审通过", "📅 待启动"),
    ("D-10", "测试报告", "功能测试、联调、压力测试、安全测试报告",
     "吴昊+褚明", "2026-06-27", "所有关键指标达标", "🔄 进行中"),
    ("D-11", "用户使用手册", "车商端操作指南、常见问题 FAQ",
     "冯晓+产品", "2026-07-10", "用户测试反馈良好", "📅 待启动"),
    ("D-12", "UI 设计稿与规范", "Figma 设计稿、Logo、图标、组件规范",
     "郑爽", "2026-05-20", "产品+设计评审通过", "✅ 已交付"),
    ("D-13", "合同模板", "标准交易合同模板(经法务审核)",
     "张伟+法务", "2026-06-05", "法务审核通过", "✅ 已交付"),
    ("D-14", "上线运营方案", "冷启动策略、推广渠道、KPI 监控",
     "冯晓+张伟", "2026-07-07", "管理层审批通过", "📅 待启动"),
    ("D-15", "监控告警配置", "Prometheus + Grafana、告警规则",
     "钱进+李明", "2026-07-03", "关键指标全覆盖", "📅 待启动"),
    ("D-16", "应急预案", "系统宕机、数据故障、安全事件应急预案",
     "钱进+李明", "2026-07-05", "管理层评审通过", "📅 待启动"),
]

r = write_table(ws8, r,
                ["编号", "交付物名称", "交付物说明",
                 "负责人", "计划交付日期", "验收标准", "状态"],
                deliverables)

# ============ Sheet 9: 技术架构 ============
ws9 = wb.create_sheet("09 技术架构")
ws9.sheet_view.showGridLines = False
for i, w in enumerate([18, 28, 30, 38, 14, 18, 12, 14, 10, 12], 1):
    ws9.column_dimensions[get_column_letter(i)].width = w

r = write_title(ws9, 1, "技术架构与基础设施", "技术选型 · 架构说明", max_col=MAX_C)

tech_backend = [
    ("后端框架", "Spring Boot 3.5.x", "核心应用框架",
     "Java 21", "社区成熟/生产可用", "Spring Boot"),
    ("持久层", "MyBatis-Plus 3.5.x", "ORM 框架、自动 CRUD、分页",
     "强大的代码生成/条件构造器", "生产可用", "MyBatis-Plus"),
    ("主数据库", "PostgreSQL 16", "业务数据存储、31 张核心表",
     "丰富数据类型/强一致性/JSON 支持", "生产可用", "PostgreSQL"),
    ("缓存层", "Redis 7.x", "车源列表缓存、热点数据、Session",
     "高性能/持久化/发布订阅", "生产可用", "Redis"),
    ("消息队列", "Apache RocketMQ 5.x", "异步消息、订单通知、任务解耦",
     "高吞吐/事务消息/顺序消息", "生产可用", "RocketMQ"),
    ("实时通信", "WebSocket + STOMP", "实时聊天、客服消息推送",
     "全双工/低延迟/浏览器原生支持", "生产可用", "Spring WebSocket"),
    ("认证鉴权", "JWT (io.jsonwebtoken 0.12.x)", "无状态 Token、权限控制",
     "无状态/跨域友好/易于扩展", "生产可用", "JJWT"),
    ("AI 集成", "OpenAI 兼容 API + LLM", "AI 行情分析、获客文案生成",
     "可切换供应商/支持流式响应", "生产可用", "OpenAI SDK"),
    ("工具库", "Hutool 5.8.x", "通用工具（日期、字符串、加密）",
     "功能齐全/中文文档友好", "生产可用", "Hutool"),
    ("接口文档", "SpringDoc OpenAPI 2.x", "自动生成 Swagger 文档",
     "实时同步代码变更", "生产可用", "SpringDoc"),
]

r = write_section_label(ws9, r, "后端技术栈", max_col=MAX_C)
r = write_table(ws9, r,
                ["组件", "技术/版本", "用途",
                 "关键特性", "成熟度", "官方/社区"],
                tech_backend, widths=[16, 28, 32, 36, 16, 22])

tech_frontend = [
    ("前端框架", "uni-app 3.x + Vue 3.x", "跨端小程序/H5 应用框架",
     "一套代码多端运行/微信小程序/H5", "生产可用", "DCloud"),
    ("状态管理", "Vuex 4.x + Pinia", "全局状态管理、用户/车源数据",
     "DevTools 支持/模块化", "生产可用", "Vuex"),
    ("UI 组件库", "uView Plus 3.x + uView UI 2.x", "UI 组件、表单、列表、图标",
     "丰富组件/主题定制/小程序优化", "生产可用", "uView"),
    ("开发工具", "HBuilderX / VS Code", "代码编辑、调试、打包发布",
     "一键运行/多端模拟器", "生产可用", "DCloud"),
    ("构建工具", "Vite 5.x", "快速开发/构建",
     "极速热更新/按需编译", "生产可用", "Vite"),
    ("样式方案", "Sass/SCSS 1.77.x", "样式预处理、变量、混合",
     "主题变量/模块化/自动前缀", "生产可用", "Sass"),
    ("网络请求", "uni.request + 封装", "后端 API 请求、拦截器、Token 管理",
     "统一错误处理/请求拦截/响应转换", "生产可用", "uni-app"),
    ("实时通信", "uni.connectSocket + STOMP", "WebSocket 聊天、实时消息推送",
     "断线重连/心跳检测/消息队列", "生产可用", "WebSocket API"),
]

r = write_section_label(ws9, r, "前端技术栈", max_col=MAX_C)
r = write_table(ws9, r,
                ["组件", "技术/版本", "用途",
                 "关键特性", "成熟度", "官方/社区"],
                tech_frontend, widths=[16, 28, 32, 36, 16, 22])

infra = [
    ("生产服务器", "腾讯云 CVM / 阿里云 ECS", "8C16G × 2 (应用主备)",
     "应用服务部署、负载均衡、弹性伸缩", "¥ 1,200/月", "生产可用"),
    ("数据库服务器", "PostgreSQL 16 (云托管)", "16C32G × 2 (主从)",
     "高可用/自动备份/监控告警", "¥ 2,400/月", "生产可用"),
    ("Redis 缓存", "云数据库 Redis 7", "4C8G 主从 + 哨兵",
     "高可用/持久化/慢查询监控", "¥ 800/月", "生产可用"),
    ("对象存储", "阿里云 OSS / 腾讯云 COS", "不限容量",
     "车源图片/合同文件/头像上传/CDN 加速", "按量付费", "生产可用"),
    ("消息队列", "阿里云 RocketMQ 5", "按需 Topic + 消费者组",
     "高吞吐/事务消息/死信队列告警", "¥ 600/月", "生产可用"),
    ("监控告警", "Prometheus + Grafana", "应用指标/系统指标/日志聚合",
     "可视化仪表盘/自定义告警规则", "开源免费", "生产可用"),
    ("日志收集", "阿里云 SLS (Simple Log Service)", "应用日志/Nginx 日志",
     "日志检索/分析/可视化", "按量付费", "生产可用"),
    ("CDN 加速", "阿里云 CDN / 腾讯云 CDN", "静态资源/图片/文件",
     "全球加速/HTTPS/缓存策略", "按量付费", "生产可用"),
    ("SSL 证书", "Let's Encrypt / 阿里云 SSL", "全站 HTTPS 加密",
     "免费证书/自动续期/浏览器信任", "免费/商用可选", "生产可用"),
    ("域名备案", "ICP 备案 + 公安备案", "国内可访问",
     "合规要求/约 10-20 工作日", "免费", "已就绪"),
]

r = write_section_label(ws9, r, "基础设施与部署", max_col=MAX_C)
r = write_table(ws9, r,
                ["组件", "基础设施", "规格配置",
                 "用途/关键特性", "预估月费用", "状态"],
                infra, widths=[16, 28, 28, 40, 18, 14])

# ============ 保存文件 ============
output_path = r"d:\ai_project\new-car-trade\5D好车-功能开发计划.xlsx"
wb.save(output_path)
print(f"Excel 文件已生成: {output_path}")
print(f"共 {len(wb.worksheets)} 个工作表")
for ws in wb.worksheets:
    max_row = ws.max_row
    print(f"  - {ws.title}: {max_row} 行数据")
